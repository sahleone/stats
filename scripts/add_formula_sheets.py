"""
add_formula_sheets.py — Adds a "Worked Example (Formulas)" sheet to each of
the 6 existing workbooks, maximising live Excel formulas.
"""

import sys
import os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))

from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Styles (matching scripts/excel_styles.py) ─────────────────────────
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='2F5496')
HEADER_FONT = Font(name='Arial', bold=True, size=12, color='FFFFFF')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=11, color='2F5496')
BODY_FONT = Font(name='Arial', size=11)
BOLD_FONT = Font(name='Arial', bold=True, size=11)
ITALIC_FONT = Font(name='Arial', italic=True, size=10)

HEADER_FILL = PatternFill('solid', fgColor='2F5496')
LIGHT_BLUE_FILL = PatternFill('solid', fgColor='D6E4F0')
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ── Helpers ────────────────────────────────────────────────────────────
def set_col_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = TITLE_FONT
    c.alignment = LEFT_WRAP
    return row + 1


def write_sub(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = SUBHEADER_FONT
    c.alignment = LEFT_WRAP
    return row + 1


def write_body(ws, row, col, text, height=None):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BODY_FONT
    c.alignment = LEFT_WRAP
    if height:
        ws.row_dimensions[row].height = height
    else:
        lines = str(text).count('\n') + 1
        ws.row_dimensions[row].height = max(30, lines * 16)
    return row + 1


def write_bold(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = BOLD_FONT
    c.alignment = LEFT_WRAP
    return row + 1


def write_italic(ws, row, col, text):
    c = ws.cell(row=row, column=col, value=text)
    c.font = ITALIC_FONT
    c.alignment = LEFT_WRAP
    return row + 1


def header_row(ws, row, start_col, headers):
    for i, h in enumerate(headers):
        c = ws.cell(row=row, column=start_col + i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
        c.border = THIN_BORDER
    return row + 1


def data_cell(ws, row, col, value, font=None, fill=None, align=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = font or BODY_FONT
    c.fill = fill or PatternFill()
    c.alignment = align or CENTER
    c.border = THIN_BORDER
    return c


def yellow_cell(ws, row, col, value):
    return data_cell(ws, row, col, value, font=BOLD_FONT, fill=YELLOW_FILL)


def formula_ref_table(ws, row, col, entries):
    """Write a 'Formula Reference' table.
    entries: list of (function, description, example)
    """
    row += 2  # blank rows
    write_sub(ws, row, col, "Excel Functions Used in This Example")
    row += 1
    row = header_row(ws, row, col, ["Function", "What It Does", "Example"])
    for func, desc, ex in entries:
        data_cell(ws, row, col, func, align=LEFT_CENTER)
        data_cell(ws, row, col + 1, desc, align=LEFT_CENTER)
        data_cell(ws, row, col + 2, ex, align=LEFT_CENTER)
        ws.row_dimensions[row].height = 32
        row += 1
    return row


# ======================================================================
# 1. CHI-SQUARE GOODNESS OF FIT
# ======================================================================
def add_gof_sheet(filepath):
    wb = load_workbook(filepath)
    if "Worked Example (Formulas)" in wb.sheetnames:
        del wb["Worked Example (Formulas)"]
    ws = wb.create_sheet("Worked Example (Formulas)")

    set_col_widths(ws, [4, 42, 18, 18, 18, 18])

    r = 1
    r = write_title(ws, r, 2, "Worked Example (Formulas): Is a Die Fair?")
    r += 1

    # Parameters
    r = write_sub(ws, r, 2, "Parameters")
    write_bold(ws, r, 2, "Significance Level (α)")
    data_cell(ws, r, 3, 0.05, fill=YELLOW_FILL)
    alpha_row = r
    r += 1
    r += 1

    # Problem
    r = write_sub(ws, r, 2, "Problem:")
    r = write_body(ws, r, 2,
        "A statistics student rolls a six-sided die 120 times and records "
        "the following results.\nTest at the α = 0.05 significance level "
        "whether the die is fair.", height=48)
    r += 1

    # Step 1
    r = write_sub(ws, r, 2, "Step 1: State the Hypotheses")
    r = write_body(ws, r, 2,
        "H₀: The die is fair (each face has probability 1/6).")
    r = write_body(ws, r, 2,
        "H₁: The die is NOT fair (at least one face has a different "
        "probability).")
    r += 1

    # Step 2: Data
    data_start = r
    r = write_sub(ws, r, 2, "Step 2: Set Up the Data")
    r = header_row(ws, r, 2, ["Face", "Observed (O)", "Expected (E)",
                               "(O − E)", "(O − E)² / E"])

    obs = [25, 17, 15, 23, 24, 16]
    data_first = r
    for i, o in enumerate(obs, 1):
        data_cell(ws, r, 2, i, fill=LIGHT_BLUE_FILL)
        data_cell(ws, r, 3, o)
        # Expected = Total / 6, formula-driven
        data_cell(ws, r, 4, f"=SUM($C${data_first}:$C${data_first+5})/COUNTA(B{data_first}:B{data_first+5})")
        data_cell(ws, r, 5, f"=C{r}-D{r}")
        data_cell(ws, r, 6, f"=(C{r}-D{r})^2/D{r}")
        r += 1
    data_last = r - 1

    # Totals row
    data_cell(ws, r, 2, "Total", font=BOLD_FONT, fill=LIGHT_BLUE_FILL)
    data_cell(ws, r, 3, f"=SUM(C{data_first}:C{data_last})", font=BOLD_FONT)
    data_cell(ws, r, 4, f"=SUM(D{data_first}:D{data_last})", font=BOLD_FONT)
    data_cell(ws, r, 6, f"=SUM(F{data_first}:F{data_last})", font=BOLD_FONT)
    totals_row = r
    r += 2

    # Step 3: Test Statistic
    r = write_sub(ws, r, 2, "Step 3: Calculate the Test Statistic")
    write_bold(ws, r, 2, "χ² = Σ (Oᵢ − Eᵢ)² / Eᵢ  =")
    chi_sq_row = r
    yellow_cell(ws, r, 3, f"=F{totals_row}")
    r += 2

    # Step 4: Degrees of Freedom
    r = write_sub(ws, r, 2, "Step 4: Degrees of Freedom")
    write_bold(ws, r, 2, "df = k − 1 =")
    df_row = r
    yellow_cell(ws, r, 3, f"=COUNTA(B{data_first}:B{data_last})-1")
    r += 2

    # Step 5: Critical Value (FORMULA!)
    r = write_sub(ws, r, 2, "Step 5: Find the Critical Value")
    write_body(ws, r, 2, "Using CHISQ.INV.RT(α, df):")
    r += 1
    write_bold(ws, r, 2, "χ²_critical =")
    crit_row = r
    yellow_cell(ws, r, 3, f"=CHISQ.INV.RT($C${alpha_row},C{df_row})")
    r += 2

    # Step 6: P-Value (FORMULA!)
    r = write_sub(ws, r, 2, "Step 6: Calculate the P-Value")
    write_bold(ws, r, 2, "p-value =")
    pval_row = r
    yellow_cell(ws, r, 3, f"=CHISQ.DIST.RT(C{chi_sq_row},C{df_row})")
    r += 2

    # Step 7: Decision
    r = write_sub(ws, r, 2, "Step 7: Make the Decision")
    write_bold(ws, r, 2, "Decision (by critical value):")
    dec1_row = r
    yellow_cell(ws, r, 3,
        f'=IF(C{chi_sq_row}>C{crit_row},'
        f'"Reject H\u2080","Fail to Reject H\u2080")')
    r += 1
    write_bold(ws, r, 2, "Decision (by p-value):")
    dec2_row = r
    yellow_cell(ws, r, 3,
        f'=IF(C{pval_row}<$C${alpha_row},'
        f'"Reject H\u2080","Fail to Reject H\u2080")')
    r += 1
    r = write_italic(ws, r, 2,
        "Both methods should agree. Students can verify by "
        "changing α in cell C3.")
    r += 1

    # Step 8: Conclusion
    r = write_sub(ws, r, 2, "Step 8: State the Conclusion")
    r = write_body(ws, r, 2,
        "At the 0.05 significance level, there is not sufficient evidence "
        "to conclude that the die is unfair. The observed frequencies are "
        "consistent with what we would expect from a fair die.", height=48)
    r += 1

    # Formula Reference
    r = formula_ref_table(ws, r, 2, [
        ("=CHISQ.INV.RT(α, df)",
         "Returns the critical value for a right-tailed χ² test",
         "=CHISQ.INV.RT(0.05, 5) → 11.070"),
        ("=CHISQ.DIST.RT(x, df)",
         "Returns the p-value (right-tail area) for a χ² statistic",
         "=CHISQ.DIST.RT(4.6, 5) → 0.467"),
        ("=COUNTA(range)",
         "Counts the number of non-empty cells in a range",
         "=COUNTA(B14:B19) → 6"),
        ("=SUM(range)",
         "Adds all numbers in a range",
         "=SUM(F14:F19) → χ² statistic"),
        ("=IF(test, true, false)",
         "Returns one value if a condition is TRUE, another if FALSE",
         '=IF(χ²>crit, "Reject H₀", "Fail to Reject H₀")'),
    ])

    wb.save(filepath)
    print(f"  ✓ Added 'Worked Example (Formulas)' to {filepath}")


# ======================================================================
# 2. CHI-SQUARE INDEPENDENCE
# ======================================================================
def add_independence_sheet(filepath):
    wb = load_workbook(filepath)
    if "Worked Example (Formulas)" in wb.sheetnames:
        del wb["Worked Example (Formulas)"]
    ws = wb.create_sheet("Worked Example (Formulas)")

    set_col_widths(ws, [4, 42, 18, 18, 18, 18])

    r = 1
    r = write_title(ws, r, 2,
        "Worked Example (Formulas): Gender and Study Preference")
    r += 1

    # Parameters
    r = write_sub(ws, r, 2, "Parameters")
    write_bold(ws, r, 2, "Significance Level (α)")
    data_cell(ws, r, 3, 0.05, fill=YELLOW_FILL)
    alpha_row = r
    r += 2

    # Problem
    r = write_sub(ws, r, 2, "Problem:")
    r = write_body(ws, r, 2,
        "A university researcher surveys 200 students to determine whether "
        "there is an association between gender (Male / Female) and "
        "preferred study method (Alone / Group / Online). "
        "Test at α = 0.05.", height=48)
    r += 1

    # Step 1: Hypotheses
    r = write_sub(ws, r, 2, "Step 1: State the Hypotheses")
    r = write_body(ws, r, 2,
        "H₀: Gender and study preference are independent.")
    r = write_body(ws, r, 2,
        "H₁: Gender and study preference are NOT independent.")
    r += 1

    # Step 2: Observed
    r = write_sub(ws, r, 2, "Step 2: Observed Frequencies")
    obs_header_row = r
    header_row(ws, r, 3, ["Alone", "Group", "Online", "Row Total"])
    r += 1
    # Male
    obs_r1 = r
    data_cell(ws, r, 2, "Male", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, 40)
    data_cell(ws, r, 4, 30)
    data_cell(ws, r, 5, 20)
    data_cell(ws, r, 6, f"=SUM(C{r}:E{r})")
    r += 1
    # Female
    obs_r2 = r
    data_cell(ws, r, 2, "Female", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, 30)
    data_cell(ws, r, 4, 50)
    data_cell(ws, r, 5, 30)
    data_cell(ws, r, 6, f"=SUM(C{r}:E{r})")
    r += 1
    # Column totals
    col_tot_row = r
    data_cell(ws, r, 2, "Col Total", font=BOLD_FONT,
              fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=SUM(C{obs_r1}:C{obs_r2})", font=BOLD_FONT)
    data_cell(ws, r, 4, f"=SUM(D{obs_r1}:D{obs_r2})", font=BOLD_FONT)
    data_cell(ws, r, 5, f"=SUM(E{obs_r1}:E{obs_r2})", font=BOLD_FONT)
    data_cell(ws, r, 6, f"=SUM(F{obs_r1}:F{obs_r2})", font=BOLD_FONT)
    grand_total_cell = f"F{col_tot_row}"
    r += 2

    # Step 3: Expected
    r = write_sub(ws, r, 2, "Step 3: Expected Frequencies")
    r = write_body(ws, r, 2,
        "Formula: E = (Row Total × Col Total) / Grand Total")
    header_row(ws, r, 3, ["Alone", "Group", "Online"])
    r += 1
    exp_r1 = r
    data_cell(ws, r, 2, "Male", fill=GREEN_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3,
        f"=F{obs_r1}*C{col_tot_row}/{grand_total_cell}", fill=GREEN_FILL)
    data_cell(ws, r, 4,
        f"=F{obs_r1}*D{col_tot_row}/{grand_total_cell}", fill=GREEN_FILL)
    data_cell(ws, r, 5,
        f"=F{obs_r1}*E{col_tot_row}/{grand_total_cell}", fill=GREEN_FILL)
    r += 1
    exp_r2 = r
    data_cell(ws, r, 2, "Female", fill=GREEN_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3,
        f"=F{obs_r2}*C{col_tot_row}/{grand_total_cell}", fill=GREEN_FILL)
    data_cell(ws, r, 4,
        f"=F{obs_r2}*D{col_tot_row}/{grand_total_cell}", fill=GREEN_FILL)
    data_cell(ws, r, 5,
        f"=F{obs_r2}*E{col_tot_row}/{grand_total_cell}", fill=GREEN_FILL)
    r += 2

    # Step 4: (O-E)^2/E
    r = write_sub(ws, r, 2, "Step 4: Compute (O − E)² / E for Each Cell")
    header_row(ws, r, 3, ["Alone", "Group", "Online"])
    r += 1
    comp_r1 = r
    data_cell(ws, r, 2, "Male", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=(C{obs_r1}-C{exp_r1})^2/C{exp_r1}")
    data_cell(ws, r, 4, f"=(D{obs_r1}-D{exp_r1})^2/D{exp_r1}")
    data_cell(ws, r, 5, f"=(E{obs_r1}-E{exp_r1})^2/E{exp_r1}")
    r += 1
    comp_r2 = r
    data_cell(ws, r, 2, "Female", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=(C{obs_r2}-C{exp_r2})^2/C{exp_r2}")
    data_cell(ws, r, 4, f"=(D{obs_r2}-D{exp_r2})^2/D{exp_r2}")
    data_cell(ws, r, 5, f"=(E{obs_r2}-E{exp_r2})^2/E{exp_r2}")
    r += 2

    # Step 5: Test Statistic
    r = write_sub(ws, r, 2, "Step 5: Calculate the Test Statistic")
    write_bold(ws, r, 2, "χ² = Σ all cells above =")
    chi_sq_row = r
    yellow_cell(ws, r, 3,
        f"=SUM(C{comp_r1}:E{comp_r1})+SUM(C{comp_r2}:E{comp_r2})")
    r += 2

    # Step 6: Degrees of Freedom
    r = write_sub(ws, r, 2, "Step 6: Degrees of Freedom")
    write_bold(ws, r, 2, "df = (r−1)(c−1) =")
    df_row = r
    yellow_cell(ws, r, 3,
        f"=(ROWS(C{obs_r1}:C{obs_r2}))*(COLUMNS(C{obs_r1}:E{obs_r1}))-"
        f"ROWS(C{obs_r1}:C{obs_r2})-COLUMNS(C{obs_r1}:E{obs_r1})+1")
    # Simpler: =(rows-1)*(cols-1) manually counted
    # Actually let me use a cleaner formula
    r += 1
    # Overwrite with cleaner formula
    yellow_cell(ws, df_row, 3,
        f"=(ROWS(C{obs_r1}:E{obs_r2})-1)*(COLUMNS(C{obs_r1}:E{obs_r2})-1)")
    r += 1

    # Step 7: Critical Value
    r = write_sub(ws, r, 2, "Step 7: Find the Critical Value")
    write_bold(ws, r, 2, "χ²_critical =")
    crit_row = r
    yellow_cell(ws, r, 3,
        f"=CHISQ.INV.RT($C${alpha_row},C{df_row})")
    r += 2

    # Step 8: P-Value
    r = write_sub(ws, r, 2, "Step 8: Calculate the P-Value")
    write_bold(ws, r, 2, "p-value (manual) =")
    pval_row = r
    yellow_cell(ws, r, 3,
        f"=CHISQ.DIST.RT(C{chi_sq_row},C{df_row})")
    r += 1
    write_bold(ws, r, 2, "p-value (CHISQ.TEST shortcut) =")
    pval2_row = r
    yellow_cell(ws, r, 3,
        f"=CHISQ.TEST(C{obs_r1}:E{obs_r2},C{exp_r1}:E{exp_r2})")
    r += 1
    r = write_italic(ws, r, 2,
        "Note: CHISQ.TEST returns the p-value directly from observed & "
        "expected ranges — a handy one-step check.")
    r += 1

    # Step 9: Decision
    r = write_sub(ws, r, 2, "Step 9: Make the Decision")
    write_bold(ws, r, 2, "Decision (by critical value):")
    yellow_cell(ws, r, 3,
        f'=IF(C{chi_sq_row}>C{crit_row},'
        f'"Reject H\u2080","Fail to Reject H\u2080")')
    r += 1
    write_bold(ws, r, 2, "Decision (by p-value):")
    yellow_cell(ws, r, 3,
        f'=IF(C{pval_row}<$C${alpha_row},'
        f'"Reject H\u2080","Fail to Reject H\u2080")')
    r += 2

    # Step 10: Conclusion
    r = write_sub(ws, r, 2, "Step 10: State the Conclusion")
    r = write_body(ws, r, 2,
        "At the 0.05 significance level, there is sufficient evidence to "
        "conclude that gender and study preference are associated. The "
        "data suggest that males and females differ in their preferred "
        "study methods.", height=48)
    r += 1

    # Formula Reference
    r = formula_ref_table(ws, r, 2, [
        ("=CHISQ.INV.RT(α, df)",
         "Returns the critical value for a right-tailed χ² test",
         "=CHISQ.INV.RT(0.05, 2) → 5.991"),
        ("=CHISQ.DIST.RT(x, df)",
         "Returns the p-value (right-tail area) for a χ² statistic",
         "=CHISQ.DIST.RT(8.33, 2) → 0.016"),
        ("=CHISQ.TEST(observed, expected)",
         "Returns the p-value directly from observed and expected ranges",
         "=CHISQ.TEST(C14:E15, C21:E22) → p-value"),
        ("=ROWS(range) / =COLUMNS(range)",
         "Count the number of rows or columns in a range",
         "=ROWS(C14:C15) → 2"),
        ("=SUM(range)",
         "Adds all numbers in a range",
         "=SUM(C14:E14) → row total"),
        ("=IF(test, true, false)",
         "Returns one value if TRUE, another if FALSE",
         '=IF(χ²>crit, "Reject H₀", "Fail to Reject H₀")'),
    ])

    wb.save(filepath)
    print(f"  ✓ Added 'Worked Example (Formulas)' to {filepath}")


# ======================================================================
# 3. ONE-WAY ANOVA
# ======================================================================
def add_oneway_anova_sheet(filepath):
    wb = load_workbook(filepath)
    if "Worked Example (Formulas)" in wb.sheetnames:
        del wb["Worked Example (Formulas)"]
    ws = wb.create_sheet("Worked Example (Formulas)")

    set_col_widths(ws, [4, 42, 18, 18, 18, 18])

    r = 1
    r = write_title(ws, r, 2,
        "Worked Example (Formulas): Teaching Methods")
    r += 1

    # Parameters
    r = write_sub(ws, r, 2, "Parameters")
    write_bold(ws, r, 2, "Significance Level (α)")
    data_cell(ws, r, 3, 0.05, fill=YELLOW_FILL)
    alpha_row = r
    r += 2

    # Problem
    r = write_sub(ws, r, 2, "Problem:")
    r = write_body(ws, r, 2,
        "A professor wants to know whether three different teaching "
        "methods lead to different exam scores. She randomly assigns "
        "15 students to three groups (5 per group). Test at α = 0.05.",
        height=48)
    r += 1

    # Step 1: Hypotheses
    r = write_sub(ws, r, 2, "Step 1: State the Hypotheses")
    r = write_body(ws, r, 2,
        "H₀: μ₁ = μ₂ = μ₃ (all teaching methods produce equal means)")
    r = write_body(ws, r, 2,
        "H₁: At least one mean differs")
    r += 1

    # Step 2: Data
    r = write_sub(ws, r, 2, "Step 2: Data")
    header_row(ws, r, 3, ["Method A", "Method B", "Method C"])
    r += 1
    data_A = [78, 85, 82, 80, 75]
    data_B = [90, 88, 92, 85, 95]
    data_C = [70, 72, 68, 74, 66]
    d_start = r
    for i in range(5):
        data_cell(ws, r, 2, f"Student {i+1}",
                  fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
        data_cell(ws, r, 3, data_A[i])
        data_cell(ws, r, 4, data_B[i])
        data_cell(ws, r, 5, data_C[i])
        r += 1
    d_end = r - 1

    # Group stats
    data_cell(ws, r, 2, "Group Mean", font=BOLD_FONT,
              fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    mean_row = r
    data_cell(ws, r, 3, f"=AVERAGE(C{d_start}:C{d_end})", font=BOLD_FONT)
    data_cell(ws, r, 4, f"=AVERAGE(D{d_start}:D{d_end})", font=BOLD_FONT)
    data_cell(ws, r, 5, f"=AVERAGE(E{d_start}:E{d_end})", font=BOLD_FONT)
    r += 1
    data_cell(ws, r, 2, "Group n", font=BOLD_FONT,
              fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    n_row = r
    data_cell(ws, r, 3, f"=COUNT(C{d_start}:C{d_end})", font=BOLD_FONT)
    data_cell(ws, r, 4, f"=COUNT(D{d_start}:D{d_end})", font=BOLD_FONT)
    data_cell(ws, r, 5, f"=COUNT(E{d_start}:E{d_end})", font=BOLD_FONT)
    r += 2

    # Step 3: Grand Mean
    r = write_sub(ws, r, 2, "Step 3: Grand Mean and Totals")
    write_bold(ws, r, 2, "Grand Mean (x̄) =")
    gm_row = r
    data_cell(ws, r, 3,
        f"=(SUM(C{d_start}:C{d_end})+SUM(D{d_start}:D{d_end})"
        f"+SUM(E{d_start}:E{d_end}))/(C{n_row}+D{n_row}+E{n_row})")
    r += 1
    write_bold(ws, r, 2, "N (total) =")
    N_row = r
    data_cell(ws, r, 3, f"=C{n_row}+D{n_row}+E{n_row}")
    r += 1
    write_bold(ws, r, 2, "k (groups) =")
    k_row = r
    data_cell(ws, r, 3, 3)
    r += 2

    # Step 4: SS_between
    r = write_sub(ws, r, 2, "Step 4: Compute SS_between")
    write_body(ws, r, 2, "SS_between = Σ nᵢ(x̄ᵢ − x̄)²")
    r += 1
    write_bold(ws, r, 2, "SS_between =")
    ssb_row = r
    data_cell(ws, r, 3,
        f"=C{n_row}*(C{mean_row}-C{gm_row})^2"
        f"+D{n_row}*(D{mean_row}-C{gm_row})^2"
        f"+E{n_row}*(E{mean_row}-C{gm_row})^2")
    r += 2

    # Step 5: SS_within
    r = write_sub(ws, r, 2, "Step 5: Compute SS_within")
    write_body(ws, r, 2, "SS_within = Σ Σ (xᵢⱼ − x̄ᵢ)²")
    r += 1
    write_bold(ws, r, 2, "SS_within =")
    ssw_row = r
    data_cell(ws, r, 3,
        f"=SUMPRODUCT((C{d_start}:C{d_end}-C{mean_row})^2)"
        f"+SUMPRODUCT((D{d_start}:D{d_end}-D{mean_row})^2)"
        f"+SUMPRODUCT((E{d_start}:E{d_end}-E{mean_row})^2)")
    r += 2

    # Step 6: ANOVA Table
    r = write_sub(ws, r, 2, "Step 6: ANOVA Table")
    anova_h = r
    header_row(ws, r, 2, ["Source", "SS", "df", "MS", "F"])
    r += 1
    # Between
    btw_row = r
    data_cell(ws, r, 2, "Between", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=C{ssb_row}")
    data_cell(ws, r, 4, f"=C{k_row}-1")  # df1
    data_cell(ws, r, 5, f"=C{r}/D{r}")  # MS_between
    data_cell(ws, r, 6, f"=E{r}/E{r+1}")  # F
    r += 1
    # Within
    wth_row = r
    data_cell(ws, r, 2, "Within", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=C{ssw_row}")
    data_cell(ws, r, 4, f"=C{N_row}-C{k_row}")  # df2
    data_cell(ws, r, 5, f"=C{r}/D{r}")  # MS_within
    r += 1
    # Total
    data_cell(ws, r, 2, "Total", font=BOLD_FONT,
              fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=C{btw_row}+C{wth_row}", font=BOLD_FONT)
    data_cell(ws, r, 4, f"=C{N_row}-1", font=BOLD_FONT)
    total_ss_row = r
    r += 2

    # Step 7: Critical Value & P-value
    r = write_sub(ws, r, 2, "Step 7: Critical Value and P-Value")
    write_bold(ws, r, 2, "df₁ (between) =")
    data_cell(ws, r, 3, f"=D{btw_row}")
    r += 1
    write_bold(ws, r, 2, "df₂ (within) =")
    data_cell(ws, r, 3, f"=D{wth_row}")
    r += 1
    write_bold(ws, r, 2, "F_critical =")
    fcrit_row = r
    yellow_cell(ws, r, 3,
        f"=F.INV.RT($C${alpha_row},D{btw_row},D{wth_row})")
    r += 1
    write_bold(ws, r, 2, "F_calc =")
    fcalc_row = r
    yellow_cell(ws, r, 3, f"=F{btw_row}")
    r += 1
    write_bold(ws, r, 2, "p-value =")
    pval_row = r
    yellow_cell(ws, r, 3,
        f"=F.DIST.RT(F{btw_row},D{btw_row},D{wth_row})")
    r += 2

    # Step 8: Decision
    r = write_sub(ws, r, 2, "Step 8: Decision")
    write_bold(ws, r, 2, "Decision (by critical value):")
    yellow_cell(ws, r, 3,
        f'=IF(C{fcalc_row}>C{fcrit_row},'
        f'"Reject H\u2080","Fail to Reject H\u2080")')
    r += 1
    write_bold(ws, r, 2, "Decision (by p-value):")
    yellow_cell(ws, r, 3,
        f'=IF(C{pval_row}<$C${alpha_row},'
        f'"Reject H\u2080","Fail to Reject H\u2080")')
    r += 2

    # Step 9: Effect Size
    r = write_sub(ws, r, 2, "Step 9: Effect Size")
    write_bold(ws, r, 2, "η² = SS_between / SS_total =")
    eta_row = r
    yellow_cell(ws, r, 3, f"=C{ssb_row}/C{total_ss_row}")
    r += 1
    r = write_italic(ws, r, 2,
        "η² ≈ 0.87 → Large effect. Teaching method explains about "
        "87% of the variability in scores.")
    r += 1

    # Step 10: Conclusion
    r = write_sub(ws, r, 2, "Step 10: Conclusion")
    r = write_body(ws, r, 2,
        "At the 0.05 significance level, there is sufficient evidence to "
        "conclude that at least one teaching method produces a different "
        "mean exam score. The effect is large (η² ≈ 0.87). A post-hoc "
        "test (e.g., Tukey's HSD) is needed to determine which specific "
        "methods differ.", height=64)
    r += 1

    # Formula Reference
    r = formula_ref_table(ws, r, 2, [
        ("=F.INV.RT(α, df1, df2)",
         "Returns the critical value for a right-tailed F test",
         "=F.INV.RT(0.05, 2, 12) → 3.885"),
        ("=F.DIST.RT(F, df1, df2)",
         "Returns the p-value (right-tail area) for an F statistic",
         "=F.DIST.RT(40.7, 2, 12) → <0.0001"),
        ("=AVERAGE(range)",
         "Returns the arithmetic mean of a range",
         "=AVERAGE(C12:C16) → group mean"),
        ("=COUNT(range)",
         "Counts the number of numeric values in a range",
         "=COUNT(C12:C16) → 5"),
        ("=SUMPRODUCT((range-mean)^2)",
         "Sum of squared deviations from the mean",
         "=SUMPRODUCT((C12:C16-C17)^2) → SS for group"),
        ("=IF(test, true, false)",
         "Returns one value if TRUE, another if FALSE",
         '=IF(F>F_crit, "Reject H₀", "Fail to Reject H₀")'),
    ])

    wb.save(filepath)
    print(f"  ✓ Added 'Worked Example (Formulas)' to {filepath}")


# ======================================================================
# 4. TWO-WAY ANOVA
# ======================================================================
def add_twoway_anova_sheet(filepath):
    wb = load_workbook(filepath)
    if "Worked Example (Formulas)" in wb.sheetnames:
        del wb["Worked Example (Formulas)"]
    ws = wb.create_sheet("Worked Example (Formulas)")

    set_col_widths(ws, [4, 42, 18, 18, 18, 18])

    r = 1
    r = write_title(ws, r, 2,
        "Worked Example (Formulas): Exercise and Diet on Weight Loss")
    r += 1

    # Parameters
    r = write_sub(ws, r, 2, "Parameters")
    write_bold(ws, r, 2, "Significance Level (α)")
    data_cell(ws, r, 3, 0.05, fill=YELLOW_FILL)
    alpha_row = r
    r += 2

    # Problem
    r = write_sub(ws, r, 2, "Problem:")
    r = write_body(ws, r, 2,
        "A researcher studies the effect of exercise type (Cardio vs "
        "Weights) and diet (Low-Carb vs Standard) on weight loss (kg) "
        "over 8 weeks. 4 participants per cell (balanced). "
        "Test all three effects at α = 0.05.", height=48)
    r += 1

    # Step 1: Data
    r = write_sub(ws, r, 2, "Step 1: Data (Weight Loss in kg)")
    header_row(ws, r, 3, ["Low-Carb", "Standard"])
    r += 1
    write_bold(ws, r, 2, "Cardio")
    r += 1
    cardio_lc_start = r
    cardio_lc = [5.2, 6.1, 5.8, 6.5]
    cardio_std = [3.5, 4.0, 3.8, 4.2]
    for i in range(4):
        data_cell(ws, r, 2, f"P{i+1}", fill=LIGHT_BLUE_FILL,
                  align=LEFT_CENTER)
        data_cell(ws, r, 3, cardio_lc[i])
        data_cell(ws, r, 4, cardio_std[i])
        r += 1
    cardio_end = r - 1

    write_bold(ws, r, 2, "Weights")
    r += 1
    weights_start = r
    weights_lc = [4.0, 4.5, 4.2, 4.8]
    weights_std = [2.5, 3.0, 2.8, 3.2]
    for i in range(4):
        data_cell(ws, r, 2, f"P{i+1}", fill=LIGHT_BLUE_FILL,
                  align=LEFT_CENTER)
        data_cell(ws, r, 3, weights_lc[i])
        data_cell(ws, r, 4, weights_std[i])
        r += 1
    weights_end = r - 1
    r += 1

    # Design params (formula-driven)
    r = write_sub(ws, r, 2, "Design Parameters")
    write_bold(ws, r, 2, "n (per cell) =")
    n_row = r
    data_cell(ws, r, 3, f"=COUNT(C{cardio_lc_start}:C{cardio_end})")
    r += 1
    write_bold(ws, r, 2, "a (levels of Exercise) =")
    a_row = r
    data_cell(ws, r, 3, 2)
    r += 1
    write_bold(ws, r, 2, "b (levels of Diet) =")
    b_row = r
    data_cell(ws, r, 3, 2)
    r += 1
    write_bold(ws, r, 2, "N (total) =")
    N_row = r
    data_cell(ws, r, 3, f"=C{n_row}*C{a_row}*C{b_row}")
    r += 2

    # Step 2: Cell Means and Marginals
    r = write_sub(ws, r, 2, "Step 2: Cell Means and Marginal Means")
    header_row(ws, r, 3, ["Low-Carb", "Standard", "Row Mean"])
    r += 1
    m_cardio = r
    data_cell(ws, r, 2, "Cardio", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3,
        f"=AVERAGE(C{cardio_lc_start}:C{cardio_end})")
    data_cell(ws, r, 4,
        f"=AVERAGE(D{cardio_lc_start}:D{cardio_end})")
    data_cell(ws, r, 5, f"=AVERAGE(C{r},D{r})")
    r += 1
    m_weights = r
    data_cell(ws, r, 2, "Weights", fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3,
        f"=AVERAGE(C{weights_start}:C{weights_end})")
    data_cell(ws, r, 4,
        f"=AVERAGE(D{weights_start}:D{weights_end})")
    data_cell(ws, r, 5, f"=AVERAGE(C{r},D{r})")
    r += 1
    m_col = r
    data_cell(ws, r, 2, "Col Mean", font=BOLD_FONT,
              fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3,
        f"=AVERAGE(C{m_cardio},C{m_weights})", font=BOLD_FONT)
    data_cell(ws, r, 4,
        f"=AVERAGE(D{m_cardio},D{m_weights})", font=BOLD_FONT)
    data_cell(ws, r, 5,
        f"=AVERAGE(C{m_col},D{m_col})", font=BOLD_FONT)
    grand_mean = f"E{m_col}"
    r += 2

    # Step 3: Sums of Squares
    r = write_sub(ws, r, 2, "Step 3: Compute Sums of Squares")

    write_bold(ws, r, 2, "SS_A (Exercise) =")
    ssA_row = r
    data_cell(ws, r, 3,
        f"=C{b_row}*C{n_row}*(E{m_cardio}-{grand_mean})^2"
        f"+C{b_row}*C{n_row}*(E{m_weights}-{grand_mean})^2")
    r += 1

    write_bold(ws, r, 2, "SS_B (Diet) =")
    ssB_row = r
    data_cell(ws, r, 3,
        f"=C{a_row}*C{n_row}*(C{m_col}-{grand_mean})^2"
        f"+C{a_row}*C{n_row}*(D{m_col}-{grand_mean})^2")
    r += 1

    write_bold(ws, r, 2, "SS_AB (Interaction) =")
    ssAB_row = r
    data_cell(ws, r, 3,
        f"=C{n_row}*((C{m_cardio}-E{m_cardio}-C{m_col}+{grand_mean})^2"
        f"+(D{m_cardio}-E{m_cardio}-D{m_col}+{grand_mean})^2"
        f"+(C{m_weights}-E{m_weights}-C{m_col}+{grand_mean})^2"
        f"+(D{m_weights}-E{m_weights}-D{m_col}+{grand_mean})^2)")
    r += 1

    write_bold(ws, r, 2, "SS_Error =")
    ssE_row = r
    data_cell(ws, r, 3,
        f"=SUMPRODUCT((C{cardio_lc_start}:C{cardio_end}"
        f"-C{m_cardio})^2)"
        f"+SUMPRODUCT((D{cardio_lc_start}:D{cardio_end}"
        f"-D{m_cardio})^2)"
        f"+SUMPRODUCT((C{weights_start}:C{weights_end}"
        f"-C{m_weights})^2)"
        f"+SUMPRODUCT((D{weights_start}:D{weights_end}"
        f"-D{m_weights})^2)")
    r += 2

    # Step 4: ANOVA Table
    r = write_sub(ws, r, 2, "Step 4: ANOVA Table")
    header_row(ws, r, 2, ["Source", "SS", "df", "MS", "F"])
    r += 1

    # Exercise (A)
    rA = r
    data_cell(ws, r, 2, "Exercise (A)", fill=LIGHT_BLUE_FILL,
              align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=C{ssA_row}")
    data_cell(ws, r, 4, f"=C{a_row}-1")  # df_A
    data_cell(ws, r, 5, f"=C{r}/D{r}")
    data_cell(ws, r, 6, f"=E{r}/E{r+3}")  # F = MS_A / MS_Error
    r += 1

    # Diet (B)
    rB = r
    data_cell(ws, r, 2, "Diet (B)", fill=LIGHT_BLUE_FILL,
              align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=C{ssB_row}")
    data_cell(ws, r, 4, f"=C{b_row}-1")  # df_B
    data_cell(ws, r, 5, f"=C{r}/D{r}")
    data_cell(ws, r, 6, f"=E{r}/E{r+2}")  # F = MS_B / MS_Error
    r += 1

    # Interaction (A×B)
    rAB = r
    data_cell(ws, r, 2, "A × B", fill=LIGHT_BLUE_FILL,
              align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=C{ssAB_row}")
    data_cell(ws, r, 4, f"=(C{a_row}-1)*(C{b_row}-1)")  # df_AB
    data_cell(ws, r, 5, f"=C{r}/D{r}")
    data_cell(ws, r, 6, f"=E{r}/E{r+1}")  # F = MS_AB / MS_Error
    r += 1

    # Error
    rErr = r
    data_cell(ws, r, 2, "Error", fill=LIGHT_BLUE_FILL,
              align=LEFT_CENTER)
    data_cell(ws, r, 3, f"=C{ssE_row}")
    data_cell(ws, r, 4,
        f"=C{a_row}*C{b_row}*(C{n_row}-1)")  # df_Error
    data_cell(ws, r, 5, f"=C{r}/D{r}")
    r += 1

    # Total
    rTot = r
    data_cell(ws, r, 2, "Total", font=BOLD_FONT,
              fill=LIGHT_BLUE_FILL, align=LEFT_CENTER)
    data_cell(ws, r, 3,
        f"=C{rA}+C{rB}+C{rAB}+C{rErr}", font=BOLD_FONT)
    data_cell(ws, r, 4, f"=C{N_row}-1", font=BOLD_FONT)
    r += 2

    # Step 5: Critical Values & P-Values
    r = write_sub(ws, r, 2,
        "Step 5: Critical Values, P-Values, and Decisions")
    r += 1

    # For each effect: F_crit, p-value, decision
    effects = [
        ("Exercise (A)", rA),
        ("Diet (B)", rB),
        ("Interaction (A×B)", rAB),
    ]
    for label, eff_row in effects:
        write_bold(ws, r, 2, f"--- {label} ---")
        r += 1
        write_bold(ws, r, 2, f"F_critical ({label}):")
        yellow_cell(ws, r, 3,
            f"=F.INV.RT($C${alpha_row},D{eff_row},D{rErr})")
        r += 1
        write_bold(ws, r, 2, f"p-value ({label}):")
        pval_eff = r
        yellow_cell(ws, r, 3,
            f"=F.DIST.RT(F{eff_row},D{eff_row},D{rErr})")
        r += 1
        write_bold(ws, r, 2, f"Decision ({label}):")
        yellow_cell(ws, r, 3,
            f'=IF(C{pval_eff}<$C${alpha_row},'
            f'"Reject H\u2080","Fail to Reject H\u2080")')
        r += 2

    # Step 6: Effect Sizes
    r = write_sub(ws, r, 2, "Step 6: Effect Sizes")
    write_bold(ws, r, 2, "η²_A (Exercise) =")
    yellow_cell(ws, r, 3, f"=C{rA}/C{rTot}")
    r += 1
    write_bold(ws, r, 2, "η²_B (Diet) =")
    yellow_cell(ws, r, 3, f"=C{rB}/C{rTot}")
    r += 1
    write_bold(ws, r, 2, "η²_AB (Interaction) =")
    yellow_cell(ws, r, 3, f"=C{rAB}/C{rTot}")
    r += 2

    # Step 7: Conclusion
    r = write_sub(ws, r, 2, "Step 7: Conclusion")
    r = write_body(ws, r, 2,
        "Both exercise type and diet type have significant effects on "
        "weight loss. Cardio produces more weight loss than weights, "
        "and low-carb diet produces more weight loss than standard. "
        "Check the interaction F-value to determine if the diet "
        "effect depends on exercise type.", height=64)
    r += 1

    # Formula Reference
    r = formula_ref_table(ws, r, 2, [
        ("=F.INV.RT(α, df1, df2)",
         "Returns the critical value for a right-tailed F test",
         "=F.INV.RT(0.05, 1, 12) → 4.747"),
        ("=F.DIST.RT(F, df1, df2)",
         "Returns the p-value (right-tail area) for an F statistic",
         "=F.DIST.RT(F_A, 1, 12) → p-value"),
        ("=AVERAGE(range)",
         "Returns the arithmetic mean of a range",
         "=AVERAGE(C10:C13) → cell mean"),
        ("=SUMPRODUCT((range-mean)^2)",
         "Sum of squared deviations from the mean",
         "=SUMPRODUCT((C10:C13-C22)^2) → SS"),
        ("=COUNT(range)",
         "Counts the number of numeric values in a range",
         "=COUNT(C10:C13) → 4"),
        ("=IF(test, true, false)",
         "Returns one value if TRUE, another if FALSE",
         '=IF(p<α, "Reject H₀", "Fail to Reject H₀")'),
    ])

    wb.save(filepath)
    print(f"  ✓ Added 'Worked Example (Formulas)' to {filepath}")


# ======================================================================
# 5. NORMAL DISTRIBUTION
# ======================================================================
def add_normal_dist_sheet(filepath):
    wb = load_workbook(filepath)
    if "Worked Example (Formulas)" in wb.sheetnames:
        del wb["Worked Example (Formulas)"]
    ws = wb.create_sheet("Worked Example (Formulas)")

    set_col_widths(ws, [4, 50, 22, 22])

    r = 1
    r = write_title(ws, r, 2,
        "Worked Example (Formulas): IQ Scores")
    r += 1

    # Parameters
    r = write_sub(ws, r, 2, "Parameters")
    write_bold(ws, r, 2, "Population Mean (μ)")
    data_cell(ws, r, 3, 100, fill=YELLOW_FILL)
    mu_row = r
    r += 1
    write_bold(ws, r, 2, "Population Std Dev (σ)")
    data_cell(ws, r, 3, 15, fill=YELLOW_FILL)
    sigma_row = r
    r += 2

    # Problem
    r = write_sub(ws, r, 2, "Problem:")
    r = write_body(ws, r, 2,
        "IQ scores are normally distributed with μ = 100 and σ = 15. "
        "That is, X ~ N(100, 15²).\n"
        "Find: (a) P(X < 120)  (b) P(85 < X < 130)  "
        "(c) The IQ score at the 90th percentile.", height=48)
    r += 1

    # ── Part (a) ──────────────────────────────────────────────
    r = write_sub(ws, r, 2, "Part (a): P(X < 120)")
    r += 1

    write_bold(ws, r, 2, "Step 1: Convert X to Z")
    r += 1
    write_body(ws, r, 2, "Z = (X − μ) / σ =")
    z_a_row = r
    yellow_cell(ws, r, 3, f"=(120-C{mu_row})/C{sigma_row}")
    r += 2

    write_bold(ws, r, 2, "Step 2: Probability using NORM.DIST")
    r += 1
    write_body(ws, r, 2, "P(X < 120) = NORM.DIST(120, μ, σ, TRUE) =")
    prob_a_row = r
    yellow_cell(ws, r, 3,
        f"=NORM.DIST(120,C{mu_row},C{sigma_row},TRUE)")
    r += 1
    write_body(ws, r, 2, "Equivalent: NORM.S.DIST(Z, TRUE) =")
    yellow_cell(ws, r, 3, f"=NORM.S.DIST(C{z_a_row},TRUE)")
    r += 1
    r = write_italic(ws, r, 2,
        "Both give the same answer. NORM.DIST uses the raw X value; "
        "NORM.S.DIST uses the Z-score.")
    r += 1
    r = write_body(ws, r, 2,
        "Interpretation: About 90.82% of people have an IQ below 120.")
    r += 1

    # ── Part (b) ──────────────────────────────────────────────
    r = write_sub(ws, r, 2, "Part (b): P(85 < X < 130)")
    r += 1

    write_bold(ws, r, 2, "Step 1: Convert Both Values to Z")
    r += 1
    write_body(ws, r, 2, "Z₁ = (85 − μ) / σ =")
    z1_row = r
    data_cell(ws, r, 3, f"=(85-C{mu_row})/C{sigma_row}")
    r += 1
    write_body(ws, r, 2, "Z₂ = (130 − μ) / σ =")
    z2_row = r
    data_cell(ws, r, 3, f"=(130-C{mu_row})/C{sigma_row}")
    r += 2

    write_bold(ws, r, 2, "Step 2: Subtract Cumulative Probabilities")
    r += 1
    write_body(ws, r, 2,
        "P(85 < X < 130) = NORM.DIST(130,μ,σ,TRUE) − "
        "NORM.DIST(85,μ,σ,TRUE) =")
    prob_b_row = r
    yellow_cell(ws, r, 3,
        f"=NORM.DIST(130,C{mu_row},C{sigma_row},TRUE)"
        f"-NORM.DIST(85,C{mu_row},C{sigma_row},TRUE)")
    r += 1
    r = write_body(ws, r, 2,
        "About 81.85% of people have an IQ between 85 and 130.")
    r += 1

    # ── Part (c) ──────────────────────────────────────────────
    r = write_sub(ws, r, 2, "Part (c): 90th Percentile")
    r += 1

    write_bold(ws, r, 2,
        "Step 1: Find Z for the 90th Percentile")
    r += 1
    write_body(ws, r, 2, "Z = NORM.S.INV(0.90) =")
    z_c_row = r
    yellow_cell(ws, r, 3, "=NORM.S.INV(0.90)")
    r += 2

    write_bold(ws, r, 2, "Step 2: Convert to X")
    r += 1
    write_body(ws, r, 2, "X = NORM.INV(0.90, μ, σ) =")
    x_c_row = r
    yellow_cell(ws, r, 3,
        f"=NORM.INV(0.90,C{mu_row},C{sigma_row})")
    r += 1
    r = write_italic(ws, r, 2,
        "NORM.INV gives the X value directly — no need to "
        "convert from Z manually.")
    r += 1
    r = write_body(ws, r, 2,
        "A person at the 90th percentile has an IQ of approximately "
        "119.22.")
    r += 2

    # Note: NORM.DIST vs NORM.S.DIST
    r = write_sub(ws, r, 2, "Note: NORM.DIST vs NORM.S.DIST")
    r = write_body(ws, r, 2,
        'NORM.DIST(x, μ, σ, TRUE) works with any normal distribution.\n'
        'NORM.S.DIST(z, TRUE) works with the standard normal '
        '(μ=0, σ=1).\n'
        'The "S" stands for "Standard". Use whichever is more '
        'convenient — they give equivalent results.', height=64)
    r += 1

    # Formula Reference
    r = formula_ref_table(ws, r, 2, [
        ("=NORM.DIST(x, μ, σ, TRUE)",
         "Cumulative probability P(X ≤ x) for a normal distribution",
         "=NORM.DIST(120, 100, 15, TRUE) → 0.9088"),
        ("=NORM.S.DIST(z, TRUE)",
         "Cumulative probability P(Z ≤ z) for the standard normal",
         "=NORM.S.DIST(1.33, TRUE) → 0.9082"),
        ("=NORM.INV(p, μ, σ)",
         "Inverse: find X such that P(X ≤ x) = p",
         "=NORM.INV(0.90, 100, 15) → 119.22"),
        ("=NORM.S.INV(p)",
         "Inverse for standard normal: find z such that P(Z ≤ z) = p",
         "=NORM.S.INV(0.90) → 1.282"),
    ])

    wb.save(filepath)
    print(f"  ✓ Added 'Worked Example (Formulas)' to {filepath}")


# ======================================================================
# 6. RANDOM VARIABLES
# ======================================================================
def add_random_vars_sheet(filepath):
    wb = load_workbook(filepath)
    if "Worked Example (Formulas)" in wb.sheetnames:
        del wb["Worked Example (Formulas)"]
    ws = wb.create_sheet("Worked Example (Formulas)")

    set_col_widths(ws, [4, 42, 18, 18, 18, 18, 18])

    r = 1
    r = write_title(ws, r, 2,
        "Worked Example (Formulas): Number of Pets per Household")
    r += 1

    # Problem
    r = write_sub(ws, r, 2, "Problem:")
    r = write_body(ws, r, 2,
        "A survey of 500 households in a town found the following "
        "distribution for the number of pets (X) per household. "
        "Find the expected value, variance, and standard deviation "
        "of X.", height=48)
    r += 1

    # Step 1: Distribution table
    r = write_sub(ws, r, 2, "Step 1: The Probability Distribution")
    header_row(ws, r, 2,
        ["x (# pets)", "P(X = x)", "x·P(X=x)", "x²",
         "x²·P(X=x)", "Cumulative F(x)"])
    r += 1
    dist_start = r
    xs = [0, 1, 2, 3, 4]
    ps = [0.30, 0.35, 0.20, 0.10, 0.05]
    for i, (x, p) in enumerate(zip(xs, ps)):
        data_cell(ws, r, 2, x, fill=LIGHT_BLUE_FILL)
        data_cell(ws, r, 3, p)
        data_cell(ws, r, 4, f"=B{r}*C{r}")       # x*P(x)
        data_cell(ws, r, 5, f"=B{r}^2")           # x²
        data_cell(ws, r, 6, f"=E{r}*C{r}")        # x²*P(x)
        # Cumulative
        data_cell(ws, r, 7, f"=SUM(C${dist_start}:C{r})")
        r += 1
    dist_end = r - 1

    # Totals
    totals_row = r
    data_cell(ws, r, 2, "Total", font=BOLD_FONT,
              fill=LIGHT_BLUE_FILL)
    data_cell(ws, r, 3, f"=SUM(C{dist_start}:C{dist_end})",
              font=BOLD_FONT)
    data_cell(ws, r, 4, f"=SUM(D{dist_start}:D{dist_end})",
              font=BOLD_FONT)
    data_cell(ws, r, 6, f"=SUM(F{dist_start}:F{dist_end})",
              font=BOLD_FONT)
    r += 2

    # Step 2: Validity check
    r = write_sub(ws, r, 2, "Step 2: Verify the Distribution is Valid")
    write_bold(ws, r, 2, "Σ P(X = xᵢ) =")
    data_cell(ws, r, 3, f"=C{totals_row}")
    r += 1
    write_bold(ws, r, 2, "Valid distribution?")
    valid_row = r
    yellow_cell(ws, r, 3,
        f'=IF(ABS(C{totals_row}-1)<0.0001,'
        f'"Valid \u2713","Invalid \u2717")')
    r += 2

    # Step 3: E(X)
    r = write_sub(ws, r, 2, "Step 3: Calculate E(X)")
    write_bold(ws, r, 2, "E(X) = Σ xᵢ · P(X = xᵢ) =")
    ex_row = r
    yellow_cell(ws, r, 3, f"=D{totals_row}")
    r += 1
    write_bold(ws, r, 2, "Alternative (SUMPRODUCT) =")
    yellow_cell(ws, r, 3,
        f"=SUMPRODUCT(B{dist_start}:B{dist_end},"
        f"C{dist_start}:C{dist_end})")
    r += 2

    # Step 4: E(X²)
    r = write_sub(ws, r, 2, "Step 4: Calculate E(X²)")
    write_bold(ws, r, 2, "E(X²) = Σ xᵢ² · P(X = xᵢ) =")
    ex2_row = r
    data_cell(ws, r, 3, f"=F{totals_row}")
    r += 2

    # Step 5: Var(X)
    r = write_sub(ws, r, 2,
        "Step 5: Calculate Var(X) Using the Shortcut Formula")
    write_bold(ws, r, 2, "Var(X) = E(X²) − [E(X)]² =")
    var_row = r
    yellow_cell(ws, r, 3,
        f"=F{totals_row}-D{totals_row}^2")
    r += 1
    write_bold(ws, r, 2, "Alternative (SUMPRODUCT) =")
    yellow_cell(ws, r, 3,
        f"=SUMPRODUCT(E{dist_start}:E{dist_end},"
        f"C{dist_start}:C{dist_end})"
        f"-SUMPRODUCT(B{dist_start}:B{dist_end},"
        f"C{dist_start}:C{dist_end})^2")
    r += 2

    # Step 6: Std Dev
    r = write_sub(ws, r, 2, "Step 6: Calculate Standard Deviation")
    write_bold(ws, r, 2, "σ = √Var(X) =")
    sd_row = r
    yellow_cell(ws, r, 3, f"=SQRT(C{var_row})")
    r += 2

    # Step 7: Interpretation
    r = write_sub(ws, r, 2, "Step 7: Interpretation")
    r = write_body(ws, r, 2,
        "On average, a household in this town has about 1.25 pets. "
        "The standard deviation is approximately 1.11, meaning a "
        "typical household's pet count is within about 1.11 pets "
        "of the mean.", height=48)
    r += 1

    # Formula Reference
    r = formula_ref_table(ws, r, 2, [
        ("=SUMPRODUCT(range1, range2)",
         "Multiplies corresponding elements and sums the results",
         "=SUMPRODUCT(B9:B13, C9:C13) → E(X)"),
        ("=SUM(range)",
         "Adds all numbers in a range",
         "=SUM(C9:C13) → checks Σ P(x) = 1"),
        ("=SQRT(x)",
         "Returns the square root of a number",
         "=SQRT(Var) → σ"),
        ("=IF(ABS(x-1)<0.0001, ...)",
         "Checks if a value equals 1 (within floating point tolerance)",
         '=IF(ABS(total-1)<0.0001, "Valid ✓", "Invalid ✗")'),
        ("=SUM(C$9:C9) [dragged down]",
         "Running cumulative sum for the CDF column F(x)",
         "Row 1: 0.30, Row 2: 0.65, Row 3: 0.85, ..."),
    ])

    wb.save(filepath)
    print(f"  ✓ Added 'Worked Example (Formulas)' to {filepath}")


# ======================================================================
# MAIN
# ======================================================================
if __name__ == "__main__":
    base = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
    os.chdir(base)

    print("Adding 'Worked Example (Formulas)' sheets...\n")

    add_gof_sheet(
        "chi_square/chi_square_goodness_of_fit.xlsx")
    add_independence_sheet(
        "chi_square/chi_square_independence.xlsx")
    add_oneway_anova_sheet(
        "one_way_anova/one_way_anova.xlsx")
    add_twoway_anova_sheet(
        "two_way_anova/two_way_anova.xlsx")
    add_normal_dist_sheet(
        "normal_distribution/normal_distribution.xlsx")
    add_random_vars_sheet(
        "random_variables/random_variables.xlsx")

    print("\nDone! All 6 files updated.")
