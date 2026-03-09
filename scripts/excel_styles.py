"""
excel_styles.py — Shared openpyxl formatting constants and helper functions.

Import this in any topic-building script to get consistent Excel styling
across all 23 topics.

Usage:
    from scripts.excel_styles import *

    wb = Workbook()
    ws = wb.active
    write_title(ws, 1, 2, "One-Sample Z-Test")
    write_subheader(ws, 3, 2, "Definition")
    write_body(ws, 4, 2, "The one-sample z-test is used when...")
    write_header_row(ws, 6, 2, 6, ["Symbol", "Meaning", "Formula", "Notes", "Example"])
    ...
"""

from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================
# FONTS
# ============================================================
TITLE_FONT = Font(name='Arial', bold=True, size=14, color='2F5496')
HEADER_FONT = Font(name='Arial', bold=True, size=12, color='FFFFFF')
SUBHEADER_FONT = Font(name='Arial', bold=True, size=11, color='2F5496')
BODY_FONT = Font(name='Arial', size=11)
BOLD_FONT = Font(name='Arial', bold=True, size=11)
ITALIC_FONT = Font(name='Arial', italic=True, size=11)
SMALL_FONT = Font(name='Arial', size=10)
SMALL_ITALIC = Font(name='Arial', italic=True, size=10)

# ============================================================
# FILLS
# ============================================================
HEADER_FILL = PatternFill('solid', fgColor='2F5496')
LIGHT_BLUE_FILL = PatternFill('solid', fgColor='D6E4F0')
YELLOW_FILL = PatternFill('solid', fgColor='FFF2CC')
GREEN_FILL = PatternFill('solid', fgColor='E2EFDA')
LIGHT_GREY_FILL = PatternFill('solid', fgColor='F2F2F2')
WHITE_FILL = PatternFill('solid', fgColor='FFFFFF')
LIGHT_RED_FILL = PatternFill('solid', fgColor='FCE4EC')
LIGHT_ORANGE_FILL = PatternFill('solid', fgColor='FFF3E0')

# ============================================================
# BORDERS
# ============================================================
THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)
BOTTOM_BORDER = Border(bottom=Side(style='thin'))
THICK_BOTTOM = Border(bottom=Side(style='medium'))

# ============================================================
# ALIGNMENTS
# ============================================================
CENTER = Alignment(horizontal='center', vertical='center', wrap_text=True)
LEFT_WRAP = Alignment(horizontal='left', vertical='top', wrap_text=True)
LEFT_CENTER = Alignment(horizontal='left', vertical='center', wrap_text=True)
RIGHT_CENTER = Alignment(horizontal='right', vertical='center')

# ============================================================
# COLOUR CONSTANTS (for python-pptx and other uses)
# ============================================================
DARK_BLUE_HEX = '2F5496'
DARK_BLUE_RGB = (47, 84, 150)


# ============================================================
# HELPER FUNCTIONS
# ============================================================

def set_col_widths(ws, widths):
    """Set column widths from a list. Index 0 = column A."""
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def write_title(ws, row, col, text):
    """Write a title cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = TITLE_FONT
    cell.alignment = LEFT_WRAP
    return row + 1


def write_subheader(ws, row, col, text):
    """Write a subheader cell (bold, dark blue)."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = SUBHEADER_FONT
    cell.alignment = LEFT_WRAP
    return row + 1


def write_body(ws, row, col, text, height=None):
    """Write a body text cell with optional row height."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BODY_FONT
    cell.alignment = LEFT_WRAP
    if height:
        ws.row_dimensions[row].height = height
    else:
        lines = text.count('\n') + 1
        ws.row_dimensions[row].height = max(30, lines * 16)
    return row + 1


def write_bold(ws, row, col, text):
    """Write a bold body cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = BOLD_FONT
    cell.alignment = LEFT_WRAP
    return row + 1


def write_italic(ws, row, col, text):
    """Write an italic body cell."""
    cell = ws.cell(row=row, column=col, value=text)
    cell.font = ITALIC_FONT
    cell.alignment = LEFT_WRAP
    return row + 1


def write_header_row(ws, row, start_col, end_col, headers):
    """Write a formatted header row (white text on dark blue fill)."""
    for c, h in enumerate(headers, start_col):
        cell = ws.cell(row=row, column=c, value=h)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER
    return row + 1


def style_data_range(ws, min_row, max_row, min_col, max_col,
                     font=None, alignment=None):
    """Apply font, alignment, and border to a rectangular range."""
    font = font or BODY_FONT
    alignment = alignment or CENTER
    for r in range(min_row, max_row + 1):
        for c in range(min_col, max_col + 1):
            cell = ws.cell(row=r, column=c)
            cell.font = font
            cell.alignment = alignment
            cell.border = THIN_BORDER


def highlight_cell(ws, row, col, value=None, fill=None):
    """Highlight a cell with yellow fill (or custom fill) and bold font."""
    cell = ws.cell(row=row, column=col)
    if value is not None:
        cell.value = value
    cell.fill = fill or YELLOW_FILL
    cell.font = BOLD_FONT
    cell.alignment = CENTER
    cell.border = THIN_BORDER


def write_see_also(ws, row, col, references):
    """
    Write a 'See Also' box at the bottom of a notes sheet.

    references: list of tuples (test_name, when_to_use)
    Example:
        write_see_also(ws, row, 2, [
            ("One-Sample T-Test", "Use when σ is unknown"),
            ("Wilcoxon Signed-Rank", "Nonparametric alternative"),
        ])
    """
    row = write_subheader(ws, row, col, "See Also (Related Tests)")
    row += 1
    headers = ["Test", "When to Use"]
    write_header_row(ws, row, col, col + 1, headers)
    row += 1
    start = row
    for test_name, description in references:
        ws.cell(row=row, column=col, value=test_name)
        ws.cell(row=row, column=col + 1, value=description)
        row += 1
    style_data_range(ws, start, row - 1, col, col + 1,
                     font=BODY_FONT, alignment=LEFT_CENTER)
    return row


def write_type_errors(ws, row, col, context_type1, context_type2):
    """
    Write context-specific Type I and Type II error interpretations.

    context_type1: str — e.g. "Concluding the mean has changed when it has not."
    context_type2: str — e.g. "Failing to detect a real change in the mean."
    """
    row = write_subheader(ws, row, col, "Type I and Type II Errors")
    row += 1
    row = write_body(ws, row, col,
        f"Type I Error (α): {context_type1}")
    row = write_body(ws, row, col,
        f"Type II Error (β): {context_type2}")
    return row


def write_notes_section(ws, row, col, items):
    """
    Write a series of labelled notes (e.g. for the Notes sheet).

    items: list of tuples (label, text)
    Returns the next available row.
    """
    for label, text in items:
        row = write_subheader(ws, row, col, label)
        row = write_body(ws, row, col, text)
        row += 1  # blank row between sections
    return row
